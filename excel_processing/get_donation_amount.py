import requests
import json
import openpyxl

HEADERS = {'content-type': 'application/json'}
API_URL = 'https://api.justgiving.com/{}/v1/fundraising/pages/{}'
APP_ID = '3533468f'
FUND_URL = 'https://www.justgiving.com/fundraising/'

excel_loc = 'excel_examples/Example.xlsx'


class DonationGetter:
    def __init__(self, app_id, headers, api_url, fund_url):
        self.app_id = app_id
        self.fund_url = fund_url
        self.headers = headers
        self.api_url = api_url

    def process_workbook(self, workbook):
        for sheet in workbook:
            if sheet.title == 'Summary':
                continue
            else:
                for row in sheet.iter_rows(min_row=2):
                    print(row)
                    row_shortname = self._get_shortname(row)
                    print(row_shortname)
                    r_code = self._check_page_exists(row_shortname)
                    print(r_code)
                    if r_code == 200:
                        donation_amount = self._get_donation_amount(row_shortname)
                        # row[12] = self._get_donation_amount(row_shortname)
                    else:
                        print("page doesn't exist")
        return workbook

    def _check_page_exists(self, shortname):
        print(self.api_url.format(self.app_id, shortname))
        head_r = requests.head(self.api_url.format(self.app_id, shortname), headers=HEADERS)
        return head_r.status_code

    def _get_donation_amount(self, shortname):
        print(self.api_url.format(self.app_id, shortname))
        get_r = requests.get(self.api_url.format(self.app_id, shortname), headers=HEADERS)
        print(get_r.status_code)
        with open('output.json', 'w') as out:
            json.dump(get_r.json(), out)

        return get_r.json()['grandTotalRaisedExcludingGiftAid']

    def _load_workbook(self):
        return openpyxl.load_workbook(self.excel_location)

    def _get_shortname(self, row):
        shortname = row[11].value.replace(self.fund_url,"")
        return shortname


def main():
    excel_location = excel_loc
    preprocessed_workbook = openpyxl.load_workbook(excel_location)
    restless_getter = DonationGetter(APP_ID, HEADERS, API_URL, FUND_URL)
    processed_workbook = restless_getter.process_workbook(preprocessed_workbook)
    processed_workbook.save('output.xlsx')


if __name__ == '__main__':
    main()
