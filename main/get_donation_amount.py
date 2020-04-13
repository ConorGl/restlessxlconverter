from tempfile import NamedTemporaryFile

import requests
import json
import openpyxl

class DonationGetter:
    def __init__(self, app_id, headers, api_url, fund_url):
        self.app_id = app_id
        self.fund_url = fund_url
        self.headers = headers
        self.api_url = api_url

    def process_workbook(self, workbook):
        for sheet in workbook:
            if sheet.title == "Raw data - registered":
                for index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                    if row[11]:
                        row_shortname = self._get_shortname(row)
                        if self._check_page_exists(row_shortname) == 200:
                            donation_amount = self._get_donation_amount(row_shortname)
                            row[12].value = donation_amount
                    else:
                        print("page doesn't exist")
        return self._save_as_stream(workbook)

    def _check_page_exists(self, shortname):
        head_r = requests.head(
            self.api_url.format(self.app_id, shortname), headers=self.headers
        )
        return head_r.status_code

    def _get_donation_amount(self, shortname):
        get_r = requests.get(
            self.api_url.format(self.app_id, shortname), headers=self.headers
        )
        with open("output.json", "w") as out:
            json.dump(get_r.json(), out)

        return get_r.json()["grandTotalRaisedExcludingGiftAid"]

    def _load_workbook(self):
        return openpyxl.load_workbook(self.excel_location)

    def _get_shortname(self, row):
        shortname = row[11].value.replace(self.fund_url, "")
        return shortname

    def _save_as_stream(self, workbook):
        with NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
            tmp.close()
        return stream